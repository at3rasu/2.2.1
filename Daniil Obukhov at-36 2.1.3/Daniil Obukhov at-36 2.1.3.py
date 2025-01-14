import csv
import os
from operator import itemgetter
from typing import List, Dict
import re

import numpy as np
import pandas as pd
import openpyxl
from matplotlib import pyplot as plt
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from jinja2 import Environment, FileSystemLoader
import pdfkit
from xlsx2html import xlsx2html


class Salary:
    def __init__(self, salary_from : str, salary_to : str, salary_currency : str):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.average_salary = (int(float(salary_from) + float(salary_to)) / 2)


class Vacancy:
    def __init__(self, vacancy: Dict[str, str]):
        self.name = vacancy["name"]
        self.salary = Salary(salary_from=vacancy["salary_from"],
                             salary_to=vacancy["salary_to"],
                             salary_currency=vacancy["salary_currency"])
        self.area_name = vacancy["area_name"]
        self.published_at = vacancy["published_at"]
        self.year = self.published_at[:4]


class DataSet:
    def __init__(self, file_name : str):
        self.file_name = file_name
        self.vacancies_objects = self.__csv_reader()

    def __csv_reader(self) -> (List[Vacancy]):
        with open(self.file_name, encoding='utf-8-sig') as file:
            file_reader = csv.reader(file)
            lines = [row for row in file_reader]
            headlines, vacancies = lines[0], lines[1:]
        result = []
        for vacancy in vacancies:
            if (len(vacancy) == len(headlines)) and (all([v != "" for v in vacancy])):
                vacancy = [" ".join(re.sub("<.*?>", "", value).replace('\n', '; ').split()) for value in vacancy]
                vacancy = {x: y for x, y in zip([r for r in headlines], [v for v in vacancy])}
                vacancy = Vacancy(vacancy)
                result.append(vacancy)
        return result


class ParamSalary:
    def __init__(self, param : str, salary: Salary):
        self.currency_to_rub = {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055
        }

        self.param = param
        self.salary = int(salary.average_salary * self.currency_to_rub[salary.salary_currency])
        self.count_vacancy = 1

    def add_salary(self, new_salary : Salary) -> None:
        self.count_vacancy += 1
        self.salary = self.salary + new_salary.average_salary * self.currency_to_rub[new_salary.salary_currency]


class Report:
    def __init__(self, profession : str, years: List[int], average_salary : List[int], average_salary_profession : List[int], count_vacancies_by_year : List[int], count_vacancies_by_year_prof : List[int], city_salary : Dict[str, int], city_vacancies : Dict[str, int], file_name):
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.city_salary = city_salary
        self.city_vacancies = city_vacancies
        self.profession = profession
        self.file_name = file_name

    def generate_excel(self) -> None:
        if not isinstance(self.file_name, str):
            raise TypeError('')
        if os.path.basename(self.file_name).split('.')[1] != "xlsx":
            raise TypeError('')
        if os.path.exists(self.file_name):
            raise FileExistsError("")
        df = [[self.years[i], self.average_salary[i], self.average_salary_profession[i], self.count_vacancies_by_year[i], self.count_vacancies_by_year_prof[i]] for i in range(len(self.years))]
        df.insert(0, ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий", f"Количество вакансий - {self.profession}"])
        df = pd.DataFrame(df, columns=None)
        cities_of_salary, salaries = [city for city in self.city_salary], [self.city_salary[city] for city in self.city_salary]
        cities_of_vacancy, vacancies = [city for city in self.city_vacancies], ['{:.2f}'.format(self.city_vacancies[city] * 100) + "%" for city in self.city_vacancies]
        df2 = [[cities_of_salary[i], salaries[i], "", cities_of_vacancy[i], vacancies[i]] for i in range(len(cities_of_salary))]
        df2.insert(0, ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        df2 = pd.DataFrame(df2, columns=None)
        with pd.ExcelWriter(self.file_name, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Статистика по годам', index=False, header=False)
            df2.to_excel(writer, sheet_name="Статистика по городам", index=False, header=False)
        wb = openpyxl.load_workbook(self.file_name)
        worksheet1 = wb["Статистика по годам"]
        worksheet2 = wb["Статистика по городам"]
        thin = Side(border_style="thin")
        self.__add_border_and_align(worksheet1, thin, len(self.years) + 2, ["A", "B", "C", "D", "E"])
        self.__add_border_and_align(worksheet2, thin, max(len(cities_of_salary) + 2, len(cities_of_vacancy) + 2), ["A", "B", "D", "E"])
        self.__make_max_column_width(worksheet1)
        self.__make_max_column_width(worksheet2)
        wb.save(self.file_name)

    def __add_border_and_align(self, worksheet : Worksheet, side : Side, count_columns : int, rows : List[str]) -> None:
        for i in range(1, count_columns):
            for row in rows:
                if i == 1:
                    worksheet[row + str(i)].alignment = Alignment(horizontal='left')
                    worksheet[row + str(i)].font = Font(bold=True)
                if worksheet[row + str(i)].internal_value != None:
                    worksheet[row + str(i)].border = Border(top=side, bottom=side, left=side, right=side)

    def __make_max_column_width(self, worksheet : Worksheet) -> None:
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value != None:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
                else:
                    dims[cell.column] = len(str(cell.value))
        for col, value in dims.items():
            worksheet.column_dimensions[get_column_letter(col)].width = value + 2


class Graphic:
    def __init__(self, profession: str, years: List[int], average_salary: List[int],
                 average_salary_profession: List[int], count_vacancies_by_year: List[int],
                 count_vacancies_by_year_prof: List[int], city_salary: Dict[str, int], city_vacancies: Dict[str, int],
                 file_name : str):
        if not isinstance(file_name, str):
            raise TypeError('')
        if os.path.basename(file_name).split('.')[1] != "png":
            raise TypeError('')
        if os.path.exists(file_name):
            raise FileExistsError("")
        self.years = years
        self.average_salary = average_salary
        self.average_salary_profession = average_salary_profession
        self.count_vacancies_by_year = count_vacancies_by_year
        self.count_vacancies_by_year_prof = count_vacancies_by_year_prof
        self.city_salary = city_salary
        self.city_vacancies = city_vacancies
        self.profession = profession
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2, figsize=(12, 8))
        self.__grouped_bar_graph(ax1, "Уровень зарплат по годам", self.average_salary, self.years,
                                 self.average_salary_profession, 'средняя з/п', f'з/п {self.profession}')
        self.__grouped_bar_graph(ax2, 'Количество вакансий по годам', self.count_vacancies_by_year, self.years,
                                 self.count_vacancies_by_year_prof, 'Количество вакансий',
                                 f'Количество вакансий {self.profession}')
        self.__horizontal_graph(ax3)
        self.__pie_graph(ax4)
        plt.tight_layout()
        plt.show()
        fig.savefig(file_name)

    def __grouped_bar_graph(self, ax, title: str, values_x: List[int], values_y: List[int], values_x2: List[int],
                            label_x: str, label_x2: str):
        ax.grid(axis='y')
        x = np.arange(len(values_y))
        width = 0.4
        ax.bar(x - width / 2, values_x, width, label=label_x)
        ax.bar(x + width / 2, values_x2, width, label=label_x2)
        ax.set_xticks(x, values_y, rotation=90)
        ax.tick_params(axis="both", labelsize=16)
        ax.set_title(title, fontweight='normal', fontsize=20)
        ax.legend(loc="upper left", fontsize=14)

    def __horizontal_graph(self, ax):
        ax.grid(axis='x')
        plt.rcdefaults()
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(10)
        city_salary = ["\n".join(city.split(" ")) for city in self.city_salary]
        ax.barh([city for city in city_salary], [self.city_salary[key] for key in self.city_salary], align='center')
        ax.invert_yaxis()
        ax.set_title('Уровень зарплат по городам', fontweight='normal',  fontsize=20)

    def __pie_graph(self, ax):
        plt.style.use('_mpl-gallery-nogrid')
        for label in (ax.get_xticklabels() + ax.get_yticklabels()):
            label.set_fontsize(16)
        vacancies = [self.city_vacancies[v] * 100 for v in self.city_vacancies]
        cities = [city for city in self.city_vacancies]
        sum_vacancies = sum(vacancies)
        if sum_vacancies != 100:
            vacancies.insert(0, 100 - sum_vacancies)
            cities.insert(0, "Другие")
        ax.set_title('Доля вакансий по городам', fontweight='normal',  fontsize=20)
        ax.pie(vacancies, labels=cities)


class PdfConverter:
    def __init__(self, graph_name : str, excel_file_name : str, profession : str):
        self.graph = graph_name
        self.excel_file = excel_file_name
        self.prof = profession

    def generate_pdf(self):
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        graph_path = os.path.abspath(self.graph)
        out_stream2 = xlsx2html('report.xlsx', sheet="Статистика по городам")
        out_stream2.seek(0)
        out_stream = xlsx2html('report.xlsx', sheet="Статистика по годам")
        out_stream.seek(0)
        pdf_template = template.render({"prof" : self.prof,
                                        "graph": graph_path,
                                        "first_table" : out_stream.read(),
                                        "second_table" : out_stream2.read()})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class InputConnect:
    def __init__(self):
        self.input_data = []
        for question in ["Введите название файла: ", "Введите название профессии: "]:
            print(question, end="")
            self.input_data.append(input())
        self.__process_data()

    def __process_data(self) -> None:
        data = DataSet(self.input_data[0]).vacancies_objects
        data_profession = [d for d in data if self.input_data[1] in d.name]
        year_salary = self.__convert_to_param_salary(data, "year")
        cities_salary = self.__convert_to_param_salary(data, "city")
        professions_year_salary = self.__add_missing_years(self.__convert_to_param_salary(data_profession, "year"), year_salary)
        city_salary = dict(sorted({x: y for x, y in zip([r.param for r in cities_salary], [int(v.salary / v.count_vacancy) for v in cities_salary])}.items(), key=itemgetter(1), reverse=True))
        city_vacancies = dict(sorted({x: y for x, y in zip([r.param for r in cities_salary], [v.count_vacancy / len(data) for v in cities_salary])}.items(), key=itemgetter(1), reverse=True))
        year_salary, year_vacancy = self.__convert_from_param_salary_to_dict(year_salary)
        professions_year_salary, professions_year_vacancies = self.__convert_from_param_salary_to_dict(professions_year_salary)
        city_salary = {x : y for x, y in zip([key for key in city_salary if city_vacancies[key] >= 0.01][:10], [city_salary[key] for key in city_salary if city_vacancies[key] >= 0.01])}
        city_vacancies = {x : y for x, y in zip([key for key in city_vacancies if city_vacancies[key] >= 0.01][:10], [float('{:.4f}'.format(city_vacancies[key])) for key in city_vacancies if city_vacancies[key] >= 0.01])}
        output_data = { "Динамика уровня зарплат по годам:" : year_salary,
                        "Динамика количества вакансий по годам:" : year_vacancy,
                        "Динамика уровня зарплат по годам для выбранной профессии:" : professions_year_salary,
                        "Динамика количества вакансий по годам для выбранной профессии:" : professions_year_vacancies,
                        "Уровень зарплат по городам (в порядке убывания):" : city_salary,
                        "Доля вакансий по городам (в порядке убывания):" : city_vacancies}
        [print(i, output_data[i]) for i in output_data]
        excel_file = "report.xlsx"
        profession = self.input_data[1]
        report = Report(profession=profession,
                        years=[i for i in year_salary],
                        average_salary=[year_salary[i] for i in year_salary],
                        average_salary_profession=[professions_year_salary[i] for i in professions_year_salary],
                        count_vacancies_by_year=[year_vacancy[i] for i in year_vacancy],
                        count_vacancies_by_year_prof=[professions_year_vacancies[i] for i in professions_year_vacancies],
                        city_salary=city_salary,
                        city_vacancies=city_vacancies,
                        file_name=excel_file)
        report.generate_excel()
        graph_name = "graph.png"
        graph = Graphic(profession=profession,
                        years=[i for i in year_salary],
                        average_salary=[year_salary[i] for i in year_salary],
                        average_salary_profession=[professions_year_salary[i] for i in professions_year_salary],
                        count_vacancies_by_year=[year_vacancy[i] for i in year_vacancy],
                        count_vacancies_by_year_prof=[professions_year_vacancies[i] for i in professions_year_vacancies],
                        city_salary=city_salary,
                        city_vacancies=city_vacancies,
                        file_name=graph_name)
        pdf = PdfConverter(graph_name=graph_name,
                           excel_file_name=excel_file,
                           profession=profession)
        pdf.generate_pdf()

    def __convert_to_param_salary(self, vacancies: List[Vacancy], comparison_param: str) -> (List[ParamSalary]):
        param_salary = {}
        for vacancy in vacancies:
            dict_comparison_params = {"year": vacancy.year, "city": vacancy.area_name}
            param = dict_comparison_params[comparison_param]
            if not param_salary.__contains__(param):
                param_salary[param] = ParamSalary(param, vacancy.salary)
            else:
                param_salary[param].add_salary(vacancy.salary)
        return [param_salary[d] for d in param_salary]

    def __convert_from_param_salary_to_dict(self, param_salary: List[ParamSalary]) -> (Dict[int, int], Dict[int, int]):
        return {x: y for x, y in zip([int(r.param) for r in param_salary],[0 if v.count_vacancy == 0 else int(v.salary / v.count_vacancy) for v in param_salary])}, {x: y for x, y in zip([int(r.param) for r in param_salary],[v.count_vacancy for v in param_salary])}

    def __add_missing_years(self, param_salary: List[ParamSalary], year_salary : (Dict[int, int], Dict[int, int])) -> List[ParamSalary]:
        years = [i.param for i in year_salary]
        s_years = [el.param for el in param_salary]
        for y in years:
            if y not in s_years:
                param_salary.insert(int(y) - int(years[0]), ParamSalary(y, Salary("0", "0", "RUR")))
                param_salary[int(y) - int(years[0])].count_vacancy = 0
        return param_salary


InputConnect()
